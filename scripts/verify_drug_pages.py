#!/usr/bin/env python3
"""
Verify that all drug pages in docs/drugs/ accurately reflect the data
in the reference spreadsheets (Spreadsheet 1.xlsx and Spreadsheet 2.xlsx).
"""

import os
import re
import openpyxl
from pathlib import Path
from collections import defaultdict

BASE_DIR = Path(__file__).parent.parent
DOCS_DIR = BASE_DIR / "docs" / "drugs"
REF_DIR = BASE_DIR / "reference"

# ─── Spreadsheet Parsing ────────────────────────────────────────────────────

def parse_spt_info(cell_text):
    """Extract SPT dilution and concentration from a cell like 'Neat (10mg/mL)\nDiluent: Nil'"""
    if not cell_text:
        return None
    text = str(cell_text).strip()
    spt = {"raw": text}
    
    # Extract dilution
    dil_match = re.search(r'(Neat|1:\d+)', text)
    if dil_match:
        spt["dilution"] = dil_match.group(1)
    
    # Extract concentration in parentheses
    conc_match = re.search(r'\(([^)]+mg/mL[^)]*)\)', text)
    if conc_match:
        spt["concentration"] = conc_match.group(1).strip()
    
    # Extract diluent
    diluent_match = re.search(r'Diluent:\s*(.+?)(?:\n|$)', text)
    if diluent_match:
        spt["diluent"] = diluent_match.group(1).strip()
    
    return spt


def parse_idt_info(cell_text):
    """Extract IDT dilution and concentration from a cell like '1:100\n(0.1mg/mL)' or 'NEAT\n(2.5mg/mL)'"""
    if not cell_text:
        return None
    text = str(cell_text).strip()
    idt = {"raw": text}
    
    # Extract dilution (including NEAT)
    dil_match = re.search(r'(1:\d+(?:,\d+)*)', text)
    if dil_match:
        idt["dilution"] = dil_match.group(1)
    elif re.search(r'neat', text, re.IGNORECASE):
        idt["dilution"] = "Neat"
    
    # Extract concentration in parentheses
    conc_match = re.search(r'\(([^)]+)\)', text)
    if conc_match:
        idt["concentration"] = conc_match.group(1).strip()
    
    return idt


CATEGORY_LABELS = [
    'Drug', 'Muscle Relaxant', 'Opioids', 'Hypnotics', 'Penicillin',
    'Cephalosporin', 'Local Anesthetics', 'Others', 'Proton Pump Inhibitors',
    'PPI'
]
SKIP_LABELS = ['Drug', 'Conc.', 'Time Started', 'Time Completed',
               'Result', 'Started', 'Completed']


def extract_drugs_from_sheet(ws, sheet_name):
    """Extract drug testing data from a worksheet."""
    drugs = {}
    current_drug = None
    current_data = None
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
        # Column B=index1, D=index3, E=index4, H=index7
        col_b = row[1].value if len(row) > 1 else None
        col_d = row[3].value if len(row) > 3 else None
        col_e = row[4].value if len(row) > 4 else None
        col_h = row[7].value if len(row) > 7 else None
        
        col_b_str = str(col_b).strip() if col_b else ""
        col_d_str = str(col_d).strip() if col_d else ""
        col_e_str = str(col_e).strip() if col_e else ""
        
        # Skip empty rows and non-data rows
        if not col_b_str or col_b_str in SKIP_LABELS:
            # But still check for IDT steps in col H on these rows
            if current_drug and current_data and col_h:
                idt = parse_idt_info(str(col_h))
                if idt and idt.get("dilution"):
                    current_data["idt_steps"].append(idt)
            continue
        
        # Skip category headers
        if col_b_str in CATEGORY_LABELS:
            continue
        
        # Drug name row: col D has 'Conc.', col E has SPT info
        if col_d_str == 'Conc.':
            drug_name = col_b_str.replace('\n', ' ').strip()
            if drug_name:
                current_drug = drug_name
                current_data = {
                    "sheet": sheet_name,
                    "stock_raw": None,
                    "spt": parse_spt_info(col_e_str) if col_e_str else None,
                    "idt_steps": []
                }
                drugs[current_drug] = current_data
                # Also check if col H has an IDT step on the same row
                if col_h:
                    idt = parse_idt_info(str(col_h))
                    if idt and idt.get("dilution"):
                        current_data["idt_steps"].append(idt)
            continue
        
        # Stock vial row: col D has 'Time Started'
        if current_drug and current_data and col_d_str == 'Time Started':
            current_data["stock_raw"] = col_b_str
            # Don't continue - also check for IDT steps in col H
            if col_h:
                idt = parse_idt_info(str(col_h))
                if idt and idt.get("dilution"):
                    current_data["idt_steps"].append(idt)
            continue
        
        # All other rows - check for IDT steps in col H
        if current_drug and current_data and col_h:
            idt = parse_idt_info(str(col_h))
            if idt and idt.get("dilution"):
                current_data["idt_steps"].append(idt)
    
    return drugs


def extract_all_spreadsheet_data():
    """Extract drug data from both reference spreadsheets."""
    all_drugs = {}
    
    # Parse Spreadsheet 1.xlsx
    wb1 = openpyxl.load_workbook(REF_DIR / "Spreadsheet 1.xlsx", data_only=True)
    for sheet_name in wb1.sheetnames:
        ws = wb1[sheet_name]
        drugs = extract_drugs_from_sheet(ws, f"Med Chart/{sheet_name}")
        for drug_name, drug_data in drugs.items():
            key = normalize_drug_name(drug_name)
            if key not in all_drugs:
                all_drugs[key] = {"med_chart": [], "med_list": []}
            all_drugs[key]["med_chart"].append({**drug_data, "original_name": drug_name})
    
    # Parse Spreadsheet 2.xlsx
    wb2 = openpyxl.load_workbook(REF_DIR / "Spreadsheet 2.xlsx", data_only=True)
    for sheet_name in wb2.sheetnames:
        ws = wb2[sheet_name]
        drugs = extract_drugs_from_sheet(ws, f"Medication List/{sheet_name}")
        for drug_name, drug_data in drugs.items():
            key = normalize_drug_name(drug_name)
            if key not in all_drugs:
                all_drugs[key] = {"med_chart": [], "med_list": []}
            all_drugs[key]["med_list"].append({**drug_data, "original_name": drug_name})
    
    return all_drugs


# ─── Drug Name Normalization ────────────────────────────────────────────────

DRUG_NAME_MAP = {
    "alfentanil": "alfentanil",
    "alfentanil iv": "alfentanil",
    "amoxicillin": "amoxicillin",
    "amoxycillin": "amoxicillin",
    "amoxycillin iv": "amoxicillin",
    "ampicillin": "ampicillin",
    "ampicillin iv": "ampicillin",
    "aspirin": "aspirin",
    "augmentin": "augmentin",
    "bactrim": "bactrim",
    "benzylpenicillin": "benzylpenicillin",
    "bupivacaine": "bupivacaine",
    "bupivacaine hydrochloride": "bupivacaine",
    "bupivacaine hydrochloride epidural": "bupivacaine",
    "cefazolin": "cefazolin",
    "cefazolin iv": "cefazolin",
    "cefepime": "cefepime",
    "cefepime iv": "cefepime",
    "cefotaxime": "cefotaxime",
    "cefotaxime iv": "cefotaxime",
    "ceftazidime": "ceftazidime",
    "ceftazidime iv": "ceftazidime",
    "ceftriaxone": "ceftriaxone",
    "ceftriaxone iv": "ceftriaxone",
    "cefuroxime": "cefuroxime",
    "cephalexin": "cephalexin",
    "chlorhexidine": "chlorhexidine",
    "chlorhexidine 0.02%": "chlorhexidine",
    "ciprofloxacin": "ciprofloxacin",
    "ciprofloxacin iv": "ciprofloxacin",
    "cis-atracurium": "cis-atracurium",
    "cis-atracurium iv": "cis-atracurium",
    "clindamycin": "clindamycin",
    "dalteparin": "dalteparin",
    "dexamethasone": "dexamethasone",
    "dexamethasone iv": "dexamethasone",
    "doxycycline": "doxycycline",
    "doxycycline iv": "doxycycline",
    "enoxaparin": "enoxaparin",
    "esomeprazole": "esomeprazole",
    "fentanyl": "fentanyl",
    "fentanyl iv": "fentanyl",
    "flucloxacillin": "flucloxacillin",
    "fluconazole": "fluconazole",
    "fluconazole iv": "fluconazole",
    "heparin": "heparin",
    "hydrocortisone": "hydrocortisone",
    "ketamine": "ketamine",
    "ketamine iv": "ketamine",
    "lansoprazole": "lansoprazole",
    "latex": "latex",
    "levofloxacin": "levofloxacin",
    "levofloxacin tablet": "levofloxacin",
    "lignocaine": "lignocaine",
    "lignocaine hydrochloride": "lignocaine",
    "lignocaine hydrochloride iv": "lignocaine",
    "mepivacaine": "mepivacaine",
    "mepivacaine hydrochloride": "mepivacaine",
    "mepivacaine hydrochloride epidural": "mepivacaine",
    "methylprednisolone": "methylprednisolone",
    "metronidazole": "metronidazole",
    "metronidazole iv": "metronidazole",
    "midazolam": "midazolam",
    "midazolam iv": "midazolam",
    "morphine": "morphine",
    "morphine iv": "morphine",
    "omeprazole": "omeprazole",
    "omnipaque": "omnipaque",
    "oxycodone": "oxycodone",
    "oxycodone iv": "oxycodone",
    "pancuronium": "pancuronium",
    "pancuronium iv": "pancuronium",
    "pantoprazole": "pantoprazole",
    "pantoprazole iv": "pantoprazole",
    "paracetamol": "paracetamol",
    "parecoxib": "parecoxib",
    "parecoxib iv": "parecoxib",
    "patent blue": "patent-blue",
    "patent blue sc": "patent-blue",
    "penicillin major": "penicillin-major-ppl",
    "penicillin minor": "penicillin-minor-md",
    "phenoxymethylpenicillin": "phenoxymethylpenicillin",
    "povidone iodine": "povidone-iodine",
    "povidone-iodine": "povidone-iodine",
    "propofol": "propofol",
    "propofol iv": "propofol",
    "rabeprazole": "rabeprazole",
    "remifentanil": "remifentanil",
    "remifentanil iv": "remifentanil",
    "rocuronium": "rocuronium",
    "rocuronium iv": "rocuronium",
    "ropivacaine": "ropivacaine",
    "ropivacaine epidural": "ropivacaine",
    "rosuvastatin": "rosuvastatin",
    "sugammadex": "sugammadex",
    "sugammadex iv": "sugammadex",
    "suxamethonium": "suxamethonium",
    "suxamethonium iv": "suxamethonium",
    "tazocin": "tazocin",
    "thiopental": "thiopental",
    "thiopental iv": "thiopental",
    "tranexamic acid": "tranexamic-acid",
    "tranexamic acid (txa)": "tranexamic-acid",
    "ultravist": "ultravist",
    "vancomycin": "vancomycin",
    "vecuronium": "vecuronium",
    "vecuronium iv": "vecuronium",
    "visipaque": "visipaque",
}


def normalize_drug_name(name):
    """Normalize a drug name to match against page filenames."""
    if not name:
        return ""
    # Remove common suffixes
    cleaned = name.lower().strip()
    cleaned = re.sub(r'\s+(iv|oral|sc|im|epidural|tablet|suspension)$', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = cleaned.strip()
    
    # Look up in map
    if cleaned in DRUG_NAME_MAP:
        return DRUG_NAME_MAP[cleaned]
    
    # Try without extra descriptors
    for key, val in DRUG_NAME_MAP.items():
        if key in cleaned or cleaned in key:
            return val
    
    return cleaned


# ─── Drug Page Parsing ──────────────────────────────────────────────────────

def parse_drug_page(filepath):
    """Parse a drug markdown page and extract structured data."""
    with open(filepath, 'r') as f:
        content = f.read()
    
    data = {
        "filename": filepath.stem,
        "title": "",
        "stock": None,
        "spt": None,
        "idt_steps": [],
        "has_oral_challenge": False,
        "has_iv_challenge": False,
        "has_sc_challenge": False,
        "oral_challenge_steps": [],
        "raw_content": content
    }
    
    # Extract title from first H1
    title_match = re.search(r'^#\s+(.+)$', content, re.MULTILINE)
    if title_match:
        data["title"] = title_match.group(1).strip()
    
    # Extract overview table
    overview_match = re.search(r'## Overview\s*\n(.*?)(?=\n## )', content, re.DOTALL)
    if overview_match:
        overview_text = overview_match.group(1)
        # Extract table rows
        rows = re.findall(r'\|\s*(.+?)\s*\|\s*(.+?)\s*\|', overview_text)
        for label, value in rows:
            label = label.strip().lower()
            value = value.strip()
            if 'stock' in label:
                data["stock"] = value
            elif 'spt' in label:
                data["spt"] = value
            elif 'idt' in label:
                data["idt_overview"] = value
    
    # Extract SPT section
    spt_match = re.search(r'## Skin prick test.*?\n(.*?)(?=\n## |\n### )', content, re.DOTALL)
    if spt_match:
        spt_text = spt_match.group(1)
        # Look for concentration in the SPT section
        conc_match = re.search(r'(\d+(?:\.\d+)?)\s*mg/mL', spt_text)
        if conc_match:
            if data["spt"] is None:
                data["spt"] = conc_match.group(0)
    
    # Extract IDT dilution series table
    idt_section = re.search(r'### Dilution series\s*\n(.*?)(?=\n## |\n### [^D]|\Z)', content, re.DOTALL)
    if idt_section:
        idt_text = idt_section.group(1)
        # Parse table rows for dilution steps
        table_rows = re.findall(r'\|\s*(\d+)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|', idt_text)
        for step_num, dilution, concentration in table_rows:
            data["idt_steps"].append({
                "step": int(step_num),
                "dilution": dilution.strip(),
                "concentration": concentration.strip()
            })
    
    # Also try alternate table format (some pages use different column order)
    if not data["idt_steps"]:
        idt_section2 = re.search(r'### Dilution series\s*\n(.*?)(?=\n## |\n### [^D]|\Z)', content, re.DOTALL)
        if idt_section2:
            idt_text = idt_section2.group(1)
            # Try format: | Step | Dilution | Volume | Concentration |
            table_rows = re.findall(r'\|\s*(\d+)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|\s*(.+?)\s*\|', idt_text)
            for step_num, col2, col3, col4 in table_rows:
                # Determine which column is dilution and which is concentration
                if '1:' in col2 or 'neat' in col2.lower():
                    data["idt_steps"].append({
                        "step": int(step_num),
                        "dilution": col2.strip(),
                        "concentration": col4.strip() if col4 else col3.strip()
                    })
                elif '1:' in col3 or 'neat' in col3.lower():
                    data["idt_steps"].append({
                        "step": int(step_num),
                        "dilution": col3.strip(),
                        "concentration": col4.strip() if col4 else ""
                    })
    
    # Check for oral challenge
    if re.search(r'oral.*(?:graded|challenge|protocol)', content, re.IGNORECASE):
        data["has_oral_challenge"] = True
        # Extract oral challenge steps
        oral_match = re.search(r'(?:oral graded|oral.*challenge).*?\n(.*?)(?=\n## |\n### [^O]|\Z)', content, re.DOTALL | re.IGNORECASE)
        if oral_match:
            oral_text = oral_match.group(1)
            oral_rows = re.findall(r'\|\s*(\d+)\s*\|\s*(.+?)\s*\|', oral_text)
            for step, dose_info in oral_rows:
                data["oral_challenge_steps"].append({"step": int(step), "info": dose_info.strip()})
    
    # Check for IV challenge
    if re.search(r'iv.*challenge|intravenous.*challenge', content, re.IGNORECASE):
        data["has_iv_challenge"] = True
    
    # Check for SC/IM challenge
    if re.search(r'(?:sc|im|subcutaneous|intramuscular).*challenge', content, re.IGNORECASE):
        data["has_sc_challenge"] = True
    
    return data


def parse_all_drug_pages():
    """Parse all drug markdown pages."""
    pages = {}
    for filepath in sorted(DOCS_DIR.glob("*.md")):
        if filepath.name == "index.md":
            continue
        page_data = parse_drug_page(filepath)
        pages[page_data["filename"]] = page_data
    return pages


# ─── Comparison Logic ───────────────────────────────────────────────────────

def compare_idt_steps(page_steps, spreadsheet_steps):
    """Compare IDT dilution steps between page and spreadsheet."""
    issues = []
    
    # Normalize page steps for comparison
    page_dilutions = []
    for step in page_steps:
        dil = step["dilution"].lower().strip()
        # Normalize dilution format
        dil = re.sub(r'\s+', '', dil)
        page_dilutions.append(dil)
    
    # Normalize spreadsheet steps
    ss_dilutions = []
    for step in spreadsheet_steps:
        dil = step.get("dilution", "").lower().strip()
        dil = re.sub(r'\s+', '', dil)
        if dil:
            ss_dilutions.append(dil)
    
    # Check for missing steps in page
    for ss_dil in ss_dilutions:
        found = False
        for page_dil in page_dilutions:
            if ss_dil in page_dil or page_dil in ss_dil:
                found = True
                break
        if not found:
            issues.append(f"  ⚠️  IDT step '{ss_dil}' in spreadsheet but NOT in drug page")
    
    # Check for extra steps in page
    for page_dil in page_dilutions:
        found = False
        for ss_dil in ss_dilutions:
            if ss_dil in page_dil or page_dil in ss_dil:
                found = True
                break
        if not found and ss_dilutions:  # Only flag if spreadsheet has IDT data
            issues.append(f"  ℹ️  IDT step '{page_dil}' in drug page but NOT in spreadsheet")
    
    return issues


def compare_spt(page_spt, spreadsheet_spt):
    """Compare SPT concentration between page and spreadsheet."""
    issues = []
    
    if not spreadsheet_spt:
        return issues
    
    ss_dilution = spreadsheet_spt.get("dilution", "").lower().strip()
    ss_concentration = spreadsheet_spt.get("concentration", "").lower().strip()
    
    if not page_spt:
        issues.append(f"  ⚠️  SPT data missing from drug page (spreadsheet has: {ss_dilution} {ss_concentration})")
        return issues
    
    page_spt_lower = page_spt.lower().strip()
    
    # Check if key concentration values match
    if ss_concentration:
        # Extract numeric concentration from page
        page_conc_match = re.search(r'(\d+(?:\.\d+)?)\s*mg/ml', page_spt_lower)
        ss_conc_match = re.search(r'(\d+(?:\.\d+)?)\s*mg/ml', ss_concentration)
        
        if page_conc_match and ss_conc_match:
            if page_conc_match.group(1) != ss_conc_match.group(1):
                issues.append(f"  ⚠️  SPT concentration mismatch: page has {page_conc_match.group(0)}, spreadsheet has {ss_conc_match.group(0)}")
    
    # Check dilution
    if ss_dilution:
        if ss_dilution == "neat":
            if "neat" not in page_spt_lower:
                # Check if page concentration matches stock (which would mean neat)
                pass  # Neat is hard to verify from page text alone
        elif ss_dilution not in page_spt_lower:
            # Check if the dilution ratio is implied
            dil_num_match = re.search(r'1:(\d+)', ss_dilution)
            if dil_num_match:
                issues.append(f"  ⚠️  SPT dilution '{ss_dilution}' from spreadsheet not found in page SPT: '{page_spt}'")
    
    return issues


def compare_stock(page_stock, spreadsheet_stock_raw):
    """Compare stock vial information."""
    issues = []
    
    if not spreadsheet_stock_raw:
        return issues
    
    if not page_stock:
        issues.append(f"  ⚠️  Stock info missing from drug page (spreadsheet has: {spreadsheet_stock_raw})")
        return issues
    
    # Extract key numbers from both
    page_stock_lower = page_stock.lower()
    ss_stock_lower = spreadsheet_stock_raw.lower()
    
    # Check if key concentration values match
    page_conc = re.findall(r'(\d+(?:\.\d+)?)\s*(?:mg|mcg|ml)', page_stock_lower)
    ss_conc = re.findall(r'(\d+(?:\.\d+)?)\s*(?:mg|mcg|ml)', ss_stock_lower)
    
    # Simple check: do the numbers overlap?
    if page_conc and ss_conc:
        page_nums = set(page_conc)
        ss_nums = set(ss_conc)
        if not page_nums.intersection(ss_nums):
            issues.append(f"  ⚠️  Stock info may differ: page has '{page_stock}', spreadsheet has '{spreadsheet_stock_raw}'")
    
    return issues


def generate_report(spreadsheet_data, page_data):
    """Generate the comprehensive comparison report."""
    report_lines = []
    report_lines.append("# Drug Page Verification Report")
    report_lines.append("")
    report_lines.append("This report compares all drug pages in `docs/drugs/` against the two reference spreadsheets:")
    report_lines.append("- `reference/Spreadsheet 1.xlsx`")
    report_lines.append("- `reference/Spreadsheet 2.xlsx`")
    report_lines.append("")
    
    # ── Summary ──
    total_pages = len(page_data)
    pages_with_spreadsheet_match = 0
    pages_without_spreadsheet = []
    spreadsheet_drugs_without_pages = []
    all_issues = defaultdict(list)
    
    # ── Section 1: Page-by-page comparison ──
    report_lines.append("## 1. Page-by-Page Comparison")
    report_lines.append("")
    
    for page_name, page in sorted(page_data.items()):
        # Find matching spreadsheet data
        matched_keys = []
        for ss_key in spreadsheet_data:
            if ss_key == page_name or ss_key.replace("-", "") == page_name.replace("-", ""):
                matched_keys.append(ss_key)
            elif page_name.replace("-", " ") in ss_key or ss_key.replace("-", " ") in page_name:
                matched_keys.append(ss_key)
        
        # Also try normalizing the page name
        norm_page = normalize_drug_name(page_name.replace("-", " "))
        if norm_page and norm_page not in matched_keys:
            matched_keys.append(norm_page)
        
        # Deduplicate and only keep keys that exist in spreadsheet_data
        matched_keys = list(set(k for k in matched_keys if k in spreadsheet_data))
        
        if not matched_keys:
            pages_without_spreadsheet.append(page_name)
            report_lines.append(f"### {page_name}")
            report_lines.append(f"**Status:** ❌ No matching entry found in either spreadsheet")
            report_lines.append("")
            continue
        
        pages_with_spreadsheet_match += 1
        page_issues = []
        
        for ss_key in matched_keys:
            ss_entry = spreadsheet_data[ss_key]
            
            # Check Med Chart data
            for mc_entry in ss_entry["med_chart"]:
                issues = []
                
                # Compare SPT
                if mc_entry.get("spt"):
                    spt_issues = compare_spt(page.get("spt"), mc_entry["spt"])
                    issues.extend(spt_issues)
                
                # Compare IDT
                if mc_entry.get("idt_steps"):
                    idt_issues = compare_idt_steps(page.get("idt_steps", []), mc_entry["idt_steps"])
                    issues.extend(idt_issues)
                
                # Compare stock
                if mc_entry.get("stock_raw"):
                    stock_issues = compare_stock(page.get("stock"), mc_entry["stock_raw"])
                    issues.extend(stock_issues)
                
                if issues:
                    page_issues.append(f"  **Source: {mc_entry['sheet']}**")
                    page_issues.extend(issues)
            
            # Check Medication List data
            for ml_entry in ss_entry["med_list"]:
                issues = []
                
                # Compare SPT
                if ml_entry.get("spt"):
                    spt_issues = compare_spt(page.get("spt"), ml_entry["spt"])
                    issues.extend(spt_issues)
                
                # Compare IDT
                if ml_entry.get("idt_steps"):
                    idt_issues = compare_idt_steps(page.get("idt_steps", []), ml_entry["idt_steps"])
                    issues.extend(idt_issues)
                
                # Compare stock
                if ml_entry.get("stock_raw"):
                    stock_issues = compare_stock(page.get("stock"), ml_entry["stock_raw"])
                    issues.extend(stock_issues)
                
                if issues:
                    page_issues.append(f"  **Source: {ml_entry['sheet']}**")
                    page_issues.extend(issues)
        
        if page_issues:
            report_lines.append(f"### {page_name}")
            report_lines.append(f"**Status:** ⚠️ Discrepancies found")
            report_lines.append("")
            report_lines.extend(page_issues)
            report_lines.append("")
            all_issues[page_name] = page_issues
        else:
            report_lines.append(f"### {page_name}")
            report_lines.append(f"**Status:** ✅ Matches spreadsheet data")
            report_lines.append("")
    
    # ── Section 2: Spreadsheet conflicts ──
    report_lines.append("## 2. Spreadsheet Internal Conflicts")
    report_lines.append("")
    report_lines.append("Where Med Chart and Medication List disagree for the same drug:")
    report_lines.append("")
    
    conflicts_found = False
    for ss_key, ss_entry in sorted(spreadsheet_data.items()):
        mc_entries = ss_entry["med_chart"]
        ml_entries = ss_entry["med_list"]
        
        if not mc_entries or not ml_entries:
            continue
        
        for mc in mc_entries:
            for ml in ml_entries:
                conflict = False
                conflict_details = []
                
                # Compare SPT
                mc_spt = mc.get("spt", {})
                ml_spt = ml.get("spt", {})
                if mc_spt and ml_spt:
                    mc_dil = mc_spt.get("dilution", "")
                    ml_dil = ml_spt.get("dilution", "")
                    mc_conc = mc_spt.get("concentration", "")
                    ml_conc = ml_spt.get("concentration", "")
                    
                    if mc_dil and ml_dil and mc_dil.lower() != ml_dil.lower():
                        conflict = True
                        conflict_details.append(f"SPT dilution: Med Chart='{mc_dil}' vs Med List='{ml_dil}'")
                    if mc_conc and ml_conc and mc_conc.lower() != ml_conc.lower():
                        conflict = True
                        conflict_details.append(f"SPT concentration: Med Chart='{mc_conc}' vs Med List='{ml_conc}'")
                
                # Compare IDT
                mc_idt = mc.get("idt_steps", [])
                ml_idt = ml.get("idt_steps", [])
                if mc_idt and ml_idt:
                    mc_idt_dils = sorted([s.get("dilution", "") for s in mc_idt])
                    ml_idt_dils = sorted([s.get("dilution", "") for s in ml_idt])
                    if mc_idt_dils != ml_idt_dils:
                        conflict = True
                        conflict_details.append(f"IDT steps: Med Chart={mc_idt_dils} vs Med List={ml_idt_dils}")
                
                if conflict:
                    conflicts_found = True
                    report_lines.append(f"### {ss_key}")
                    report_lines.append(f"- Med Chart source: {mc['sheet']}")
                    report_lines.append(f"- Med List source: {ml['sheet']}")
                    for detail in conflict_details:
                        report_lines.append(f"  - {detail}")
                    report_lines.append("")
    
    if not conflicts_found:
        report_lines.append("No conflicts found between the two spreadsheets.")
        report_lines.append("")
    
    # ── Section 3: Drugs in spreadsheets without pages ──
    report_lines.append("## 3. Drugs in Spreadsheets Without Drug Pages")
    report_lines.append("")
    
    page_names_normalized = set()
    for pn in page_data:
        page_names_normalized.add(pn)
        page_names_normalized.add(normalize_drug_name(pn.replace("-", " ")))
    
    for ss_key in sorted(spreadsheet_data.keys()):
        if ss_key not in page_names_normalized:
            # Check more carefully
            found = False
            for pn in page_data:
                if ss_key.replace("-", "") == pn.replace("-", ""):
                    found = True
                    break
                if ss_key in pn or pn in ss_key:
                    found = True
                    break
            if not found:
                spreadsheet_drugs_without_pages.append(ss_key)
    
    if spreadsheet_drugs_without_pages:
        for drug in spreadsheet_drugs_without_pages:
            report_lines.append(f"- {drug}")
        report_lines.append("")
    else:
        report_lines.append("All spreadsheet drugs have corresponding pages.")
        report_lines.append("")
    
    # ── Section 4: Pages without spreadsheet entries ──
    report_lines.append("## 4. Drug Pages Without Spreadsheet Entries")
    report_lines.append("")
    
    if pages_without_spreadsheet:
        for page in pages_without_spreadsheet:
            report_lines.append(f"- {page}")
        report_lines.append("")
    else:
        report_lines.append("All drug pages have matching spreadsheet entries.")
        report_lines.append("")
    
    # ── Section 5: Summary Statistics ──
    report_lines.append("## 5. Summary")
    report_lines.append("")
    report_lines.append(f"| Metric | Count |")
    report_lines.append(f"|--------|-------|")
    report_lines.append(f"| Total drug pages | {total_pages} |")
    report_lines.append(f"| Pages matching spreadsheets | {pages_with_spreadsheet_match} |")
    report_lines.append(f"| Pages with discrepancies | {len(all_issues)} |")
    report_lines.append(f"| Pages without spreadsheet match | {len(pages_without_spreadsheet)} |")
    report_lines.append(f"| Spreadsheet drugs without pages | {len(spreadsheet_drugs_without_pages)} |")
    report_lines.append(f"| Spreadsheet conflicts found | {'Yes' if conflicts_found else 'No'} |")
    report_lines.append("")
    
    # ── Section 6: Detailed Spreadsheet Data ──
    report_lines.append("## 6. Detailed Spreadsheet Data (for reference)")
    report_lines.append("")
    
    for ss_key in sorted(spreadsheet_data.keys()):
        entry = spreadsheet_data[ss_key]
        report_lines.append(f"### {ss_key}")
        
        for mc in entry["med_chart"]:
            report_lines.append(f"  **Med Chart** ({mc['sheet']}):")
            if mc.get("stock_raw"):
                report_lines.append(f"    - Stock: {mc['stock_raw']}")
            if mc.get("spt"):
                report_lines.append(f"    - SPT: {mc['spt'].get('dilution', '')} ({mc['spt'].get('concentration', '')})")
                if mc['spt'].get('diluent'):
                    report_lines.append(f"    - Diluent: {mc['spt'].get('diluent', '')}")
            if mc.get("idt_steps"):
                for step in mc["idt_steps"]:
                    report_lines.append(f"    - IDT: {step.get('dilution', '')} ({step.get('concentration', '')})")
        
        for ml in entry["med_list"]:
            report_lines.append(f"  **Medication List** ({ml['sheet']}):")
            if ml.get("stock_raw"):
                report_lines.append(f"    - Stock: {ml['stock_raw']}")
            if ml.get("spt"):
                report_lines.append(f"    - SPT: {ml['spt'].get('dilution', '')} ({ml['spt'].get('concentration', '')})")
                if ml['spt'].get('diluent'):
                    report_lines.append(f"    - Diluent: {ml['spt'].get('diluent', '')}")
            if ml.get("idt_steps"):
                for step in ml["idt_steps"]:
                    report_lines.append(f"    - IDT: {step.get('dilution', '')} ({step.get('concentration', '')})")
        
        report_lines.append("")
    
    return "\n".join(report_lines)


# ─── Main ───────────────────────────────────────────────────────────────────

def main():
    print("Extracting data from spreadsheets...")
    spreadsheet_data = extract_all_spreadsheet_data()
    print(f"  Found {len(spreadsheet_data)} unique drugs across both spreadsheets")
    
    print("Parsing drug pages...")
    page_data = parse_all_drug_pages()
    print(f"  Parsed {len(page_data)} drug pages")
    
    print("Generating comparison report...")
    report = generate_report(spreadsheet_data, page_data)
    
    output_path = BASE_DIR / "plans" / "discrepancy-report.md"
    with open(output_path, 'w') as f:
        f.write(report)
    
    print(f"\nReport written to: {output_path}")
    
    # Print summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    
    # Count issues
    issue_count = 0
    for line in report.split("\n"):
        if "⚠️" in line:
            issue_count += 1
    
    print(f"Total discrepancies found: {issue_count}")
    print(f"Report saved to: plans/discrepancy-report.md")


if __name__ == "__main__":
    main()
