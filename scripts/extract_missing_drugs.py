#!/usr/bin/env python3
"""
Extract complete SPT/IDT/stock data for all drugs that exist in the reference
spreadsheets but do NOT have corresponding pages in docs/drugs/.
"""

import os
import re
import openpyxl
from pathlib import Path
from collections import defaultdict

BASE_DIR = Path(__file__).parent.parent
DOCS_DIR = BASE_DIR / "docs" / "drugs"
REF_DIR = BASE_DIR / "reference"
OUTPUT_FILE = BASE_DIR / "plans" / "missing-drugs-data.md"

# ─── Spreadsheet Parsing ────────────────────────────────────────────────────

def parse_spt_info(cell_text):
    """Extract SPT dilution and concentration from a cell."""
    if not cell_text:
        return None
    text = str(cell_text).strip()
    spt = {"raw": text}

    dil_match = re.search(r'(Neat|1:\d+)', text)
    if dil_match:
        spt["dilution"] = dil_match.group(1)

    conc_match = re.search(r'\(([^)]+mg/mL[^)]*)\)', text)
    if conc_match:
        spt["concentration"] = conc_match.group(1).strip()

    diluent_match = re.search(r'Diluent:\s*(.+?)(?:\n|$)', text)
    if diluent_match:
        spt["diluent"] = diluent_match.group(1).strip()

    return spt


def parse_idt_info(cell_text):
    """Extract IDT dilution and concentration from a cell."""
    if not cell_text:
        return None
    text = str(cell_text).strip()
    idt = {"raw": text}

    dil_match = re.search(r'(1:\d+(?:,\d+)*)', text)
    if dil_match:
        idt["dilution"] = dil_match.group(1)
    elif re.search(r'neat', text, re.IGNORECASE):
        idt["dilution"] = "Neat"

    conc_match = re.search(r'\(([^)]+)\)', text)
    if conc_match:
        idt["concentration"] = conc_match.group(1).strip()

    return idt


CATEGORY_LABELS = [
    'Drug', 'Muscle Relaxant', 'Opioids', 'Hypnotics', 'Penicillin',
    'Cephalosporin', 'Local Anesthetics', 'Others', 'Proton Pump Inhibitors',
    'PPI', 'Antibiotics', 'NSAIDs', 'Opioid', 'Antihistamine',
    'Antifungal', 'Contrast Media', 'Muscle Relaxants', 'Hypnotic',
    'Antibiotic', 'NSAID', 'Dye', 'Dyes', 'Latex', 'Other',
    'Local Anaesthetics', 'Local Anaesthetic', 'Opiate',
    'Antibiotics (Non Penicillin)', 'Antibiotics (Penicillin)',
    'Antibiotics (Cephalosporin)', 'Antibiotics (Others)',
    'Antibiotics (Macrolide)', 'Antibiotics (Aminoglycoside)',
    'Antibiotics (Quinolone)', 'Antibiotics (Lincosamide)',
    'Antibiotics (Glycopeptide)', 'Antibiotics (Carbapenem)',
    'Antibiotics (Nitroimidazole)', 'Antibiotics (Tetracycline)',
    'Antibiotics (Sulfonamide)', 'Antibiotics (Oxazolidinone)',
    'Antibiotics (Monobactam)',
    'Hormonal', 'Hormone', 'Contraceptives', 'Contraceptive',
    'Corticosteroid', 'Corticosteroids', 'Steroid', 'Steroids',
    'Antiemetic', 'Antiemetics', 'Antipsychotic',
    'Antidiabetic', 'Insulin', 'Insulins',
    'Anticholinesterase', 'Neuromuscular',
    'Diagnostic', 'Diagnostics', 'Radiocontrast',
    'Proton pump inhibitor',
]

SKIP_LABELS = ['Drug', 'Conc.', 'Time Started', 'Time Completed',
               'Result', 'Started', 'Completed', 'Time', 'Dose',
               'Volume', 'Notes', 'Note', 'Reaction', 'Observation',
               'BP', 'HR', 'SpO2', 'RR', 'Temp', 'Patient',
               'Date', 'Ward', 'Allergy', 'Allergies']

# Sheets to skip entirely (not SPT/IDT data)
SKIP_SHEETS = {
    'Click This List to find Drug',
    'For OGC',
    'Desensitisation',
    'For IV Challenge',
    'For SC or IM Challenge',
    'For Oral Challenge',
    'OGC',
    'IV Challenge',
    'SC Challenge',
    'IM Challenge',
    'Oral Challenge',
    'Desensitization',
}


def extract_drugs_from_sheet(ws, sheet_name):
    """Extract drug testing data from a worksheet."""
    drugs = {}
    current_drug = None
    current_data = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
        # Column B=index1, D=index3, E=index4, H=index7
        col_b = row[1].value if len(row) > 1 else None
        col_d = row[3].value if len(row) > 3 else None
        col_e = row[4].value if len(row) > 4 else None
        col_h = row[7].value if len(row) > 7 else None

        col_b_str = str(col_b).strip() if col_b else ""
        col_d_str = str(col_d).strip() if col_d else ""
        col_e_str = str(col_e).strip() if col_e else ""

        # Skip empty rows and non-data rows
        if not col_b_str or col_b_str in SKIP_LABELS:
            if current_drug and current_data and col_h:
                idt = parse_idt_info(str(col_h))
                if idt and idt.get("dilution"):
                    current_data["idt_steps"].append(idt)
            continue

        # Skip category headers
        if col_b_str in CATEGORY_LABELS:
            continue

        # Drug name row: col D has 'Conc.', col E has SPT info
        if col_d_str == 'Conc.':
            drug_name = col_b_str.replace('\n', ' ').strip()
            if drug_name:
                current_drug = drug_name
                current_data = {
                    "sheet": sheet_name,
                    "stock_raw": None,
                    "spt": parse_spt_info(col_e_str) if col_e_str else None,
                    "idt_steps": []
                }
                drugs[current_drug] = current_data
                if col_h:
                    idt = parse_idt_info(str(col_h))
                    if idt and idt.get("dilution"):
                        current_data["idt_steps"].append(idt)
            continue

        # Stock vial row: col D has 'Time Started'
        if current_drug and current_data and col_d_str == 'Time Started':
            current_data["stock_raw"] = col_b_str
            if col_h:
                idt = parse_idt_info(str(col_h))
                if idt and idt.get("dilution"):
                    current_data["idt_steps"].append(idt)
            continue

        # All other rows - check for IDT steps in col H
        if current_drug and current_data and col_h:
            idt = parse_idt_info(str(col_h))
            if idt and idt.get("dilution"):
                current_data["idt_steps"].append(idt)

    return drugs


# ─── Drug Name Normalization ────────────────────────────────────────────────

DRUG_NAME_MAP = {
    "alfentanil": "alfentanil",
    "alfentanil iv": "alfentanil",
    "amoxicillin": "amoxicillin",
    "amoxycillin": "amoxicillin",
    "amoxycillin iv": "amoxicillin",
    "ampicillin": "ampicillin",
    "ampicillin iv": "ampicillin",
    "aspirin": "aspirin",
    "augmentin": "augmentin",
    "bactrim": "bactrim",
    "benzylpenicillin": "benzylpenicillin",
    "bupivacaine": "bupivacaine",
    "bupivacaine hydrochloride": "bupivacaine",
    "bupivacaine hydrochloride epidural": "bupivacaine",
    "cefazolin": "cefazolin",
    "cefazolin iv": "cefazolin",
    "cefepime": "cefepime",
    "cefepime iv": "cefepime",
    "cefotaxime": "cefotaxime",
    "cefotaxime iv": "cefotaxime",
    "ceftazidime": "ceftazidime",
    "ceftazidime iv": "ceftazidime",
    "ceftriaxone": "ceftriaxone",
    "ceftriaxone iv": "ceftriaxone",
    "cefuroxime": "cefuroxime",
    "cephalexin": "cephalexin",
    "chlorhexidine": "chlorhexidine",
    "chlorhexidine 0.02%": "chlorhexidine",
    "ciprofloxacin": "ciprofloxacin",
    "ciprofloxacin iv": "ciprofloxacin",
    "cis-atracurium": "cis-atracurium",
    "cis-atracurium iv": "cis-atracurium",
    "clindamycin": "clindamycin",
    "dalteparin": "dalteparin",
    "dexamethasone": "dexamethasone",
    "dexamethasone iv": "dexamethasone",
    "doxycycline": "doxycycline",
    "doxycycline iv": "doxycycline",
    "enoxaparin": "enoxaparin",
    "esomeprazole": "esomeprazole",
    "fentanyl": "fentanyl",
    "fentanyl iv": "fentanyl",
    "flucloxacillin": "flucloxacillin",
    "fluconazole": "fluconazole",
    "fluconazole iv": "fluconazole",
    "heparin": "heparin",
    "hydrocortisone": "hydrocortisone",
    "ketamine": "ketamine",
    "ketamine iv": "ketamine",
    "lansoprazole": "lansoprazole",
    "latex": "latex",
    "levofloxacin": "levofloxacin",
    "levofloxacin tablet": "levofloxacin",
    "lignocaine": "lignocaine",
    "lignocaine hydrochloride": "lignocaine",
    "lignocaine hydrochloride iv": "lignocaine",
    "mepivacaine": "mepivacaine",
    "mepivacaine hydrochloride": "mepivacaine",
    "mepivacaine hydrochloride epidural": "mepivacaine",
    "methylprednisolone": "methylprednisolone",
    "metronidazole": "metronidazole",
    "metronidazole iv": "metronidazole",
    "midazolam": "midazolam",
    "midazolam iv": "midazolam",
    "morphine": "morphine",
    "morphine iv": "morphine",
    "omeprazole": "omeprazole",
    "omnipaque": "omnipaque",
    "oxycodone": "oxycodone",
    "oxycodone iv": "oxycodone",
    "pancuronium": "pancuronium",
    "pancuronium iv": "pancuronium",
    "pantoprazole": "pantoprazole",
    "pantoprazole iv": "pantoprazole",
    "paracetamol": "paracetamol",
    "parecoxib": "parecoxib",
    "parecoxib iv": "parecoxib",
    "patent blue": "patent-blue",
    "patent blue sc": "patent-blue",
    "penicillin major": "penicillin-major-ppl",
    "penicillin minor": "penicillin-minor-md",
    "phenoxymethylpenicillin": "phenoxymethylpenicillin",
    "povidone iodine": "povidone-iodine",
    "povidone-iodine": "povidone-iodine",
    "propofol": "propofol",
    "propofol iv": "propofol",
    "rabeprazole": "rabeprazole",
    "remifentanil": "remifentanil",
    "remifentanil iv": "remifentanil",
    "rocuronium": "rocuronium",
    "rocuronium iv": "rocuronium",
    "ropivacaine": "ropivacaine",
    "ropivacaine epidural": "ropivacaine",
    "rosuvastatin": "rosuvastatin",
    "sugammadex": "sugammadex",
    "sugammadex iv": "sugammadex",
    "suxamethonium": "suxamethonium",
    "suxamethonium iv": "suxamethonium",
    "tazocin": "tazocin",
    "thiopental": "thiopental",
    "thiopental iv": "thiopental",
    "tranexamic acid": "tranexamic-acid",
    "tranexamic acid (txa)": "tranexamic-acid",
    "ultravist": "ultravist",
    "vancomycin": "vancomycin",
    "vecuronium": "vecuronium",
    "vecuronium iv": "vecuronium",
    "visipaque": "visipaque",
    # Missing drugs - map to likely page names
    "actrapid": "actrapid",
    "actrapid sc": "actrapid",
    "actrapid s/c": "actrapid",
    "azithromycin": "azithromycin",
    "azithromycin iv": "azithromycin",
    "betamethasone": "betamethasone",
    "betamethasone iv": "betamethasone",
    "celestone chronodose": "celestone-chronodose",
    "droperidol": "droperidol",
    "droperidol iv": "droperidol",
    "gentamicin": "gentamicin",
    "gentamicin iv": "gentamicin",
    "glycopyrronium": "glycopyrronium",
    "glycopyrronium iv": "glycopyrronium",
    "granisetron": "granisetron",
    "granisetron iv": "granisetron",
    "humulin nph": "humulin-nph",
    "humulin nph sc": "humulin-nph",
    "humulin nph s/c": "humulin-nph",
    "humulin r": "humulin-r",
    "humulin r sc": "humulin-r",
    "humulin r s/c": "humulin-r",
    "medroxyprogesterone": "medroxyprogesterone",
    "metoclopramide": "metoclopramide",
    "metoclopramide iv": "metoclopramide",
    "neostigmine": "neostigmine",
    "neostigmine iv": "neostigmine",
    "neostigmine inj": "neostigmine",
    "novorapid": "novorapid",
    "novorapid sc": "novorapid",
    "novorapid s/c": "novorapid",
    "ondansetron": "ondansetron",
    "ondansetron iv": "ondansetron",
    "optisulin": "optisulin",
    "optisulin sc": "optisulin",
    "optisulin s/c": "optisulin",
    "protamine": "protamine",
    "protamine iv": "protamine",
    "protaphane": "protaphane",
    "protaphane sc": "protaphane",
    "protaphane s/c": "protaphane",
    "tramadol": "tramadol",
    "tramadol iv": "tramadol",
    "triamcinolone": "triamcinolone",
    "triamcinolone inj": "triamcinolone",
    "urografin": "urografin",
}


def normalize_drug_name(name):
    """Normalize a drug name to match against page filenames."""
    if not name:
        return ""
    cleaned = name.lower().strip()
    # Remove common suffixes
    cleaned = re.sub(r'\s+(iv|oral|sc|im|epidural|tablet|suspension|inj|s/c)$', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    cleaned = cleaned.strip()

    if cleaned in DRUG_NAME_MAP:
        return DRUG_NAME_MAP[cleaned]

    # Try without extra descriptors
    for key, val in DRUG_NAME_MAP.items():
        if key in cleaned or cleaned in key:
            return val

    return cleaned


def get_existing_pages():
    """Get set of existing drug page names (without .md extension)."""
    pages = set()
    for filepath in sorted(DOCS_DIR.glob("*.md")):
        if filepath.name == "index.md":
            continue
        pages.add(filepath.stem)
    return pages


def extract_all_spreadsheet_data():
    """Extract drug data from both reference spreadsheets."""
    all_drugs = {}

    # Parse Spreadsheet 1.xlsx
    wb1_path = REF_DIR / "Spreadsheet 1.xlsx"
    if wb1_path.exists():
        wb1 = openpyxl.load_workbook(wb1_path, data_only=True)
        print(f"Spreadsheet 1.xlsx sheets: {wb1.sheetnames}")
        for sheet_name in wb1.sheetnames:
            # Skip non-SPT/IDT sheets
            skip = False
            for skip_pattern in SKIP_SHEETS:
                if skip_pattern.lower() in sheet_name.lower():
                    skip = True
                    break
            if skip:
                print(f"  Skipping sheet: {sheet_name}")
                continue

            ws = wb1[sheet_name]
            drugs = extract_drugs_from_sheet(ws, f"Med Chart/{sheet_name}")
            for drug_name, drug_data in drugs.items():
                key = normalize_drug_name(drug_name)
                if key not in all_drugs:
                    all_drugs[key] = []
                all_drugs[key].append({**drug_data, "original_name": drug_name, "workbook": "Spreadsheet 1.xlsx"})
    else:
        print(f"WARNING: {wb1_path} not found!")

    # Parse Spreadsheet 2.xlsx
    wb2_path = REF_DIR / "Spreadsheet 2.xlsx"
    if wb2_path.exists():
        wb2 = openpyxl.load_workbook(wb2_path, data_only=True)
        print(f"Spreadsheet 2.xlsx sheets: {wb2.sheetnames}")
        for sheet_name in wb2.sheetnames:
            skip = False
            for skip_pattern in SKIP_SHEETS:
                if skip_pattern.lower() in sheet_name.lower():
                    skip = True
                    break
            if skip:
                print(f"  Skipping sheet: {sheet_name}")
                continue

            ws = wb2[sheet_name]
            drugs = extract_drugs_from_sheet(ws, f"Medication List/{sheet_name}")
            for drug_name, drug_data in drugs.items():
                key = normalize_drug_name(drug_name)
                if key not in all_drugs:
                    all_drugs[key] = []
                all_drugs[key].append({**drug_data, "original_name": drug_name, "workbook": "Spreadsheet 2.xlsx"})
    else:
        print(f"WARNING: {wb2_path} not found!")

    return all_drugs


def format_drug_entry(key, entries):
    """Format a single missing drug's data as markdown."""
    lines = []
    lines.append(f"### {entries[0]['original_name']}")
    lines.append(f"**Normalized name:** `{key}`")
    lines.append("")

    for entry in entries:
        lines.append(f"**Source:** `{entry['workbook']}` → Sheet: `{entry['sheet']}`")
        lines.append("")

        if entry.get("stock_raw"):
            lines.append(f"- **Stock vial:** {entry['stock_raw']}")
        else:
            lines.append("- **Stock vial:** *(not found)*")

        spt = entry.get("spt")
        if spt:
            lines.append(f"- **SPT dilution:** {spt.get('dilution', 'N/A')}")
            lines.append(f"- **SPT concentration:** {spt.get('concentration', 'N/A')}")
            if spt.get("diluent"):
                lines.append(f"- **Diluent:** {spt['diluent']}")
            lines.append(f"- **SPT raw:** `{spt.get('raw', '')}`")
        else:
            lines.append("- **SPT:** *(no data)*")

        idt_steps = entry.get("idt_steps", [])
        if idt_steps:
            lines.append(f"- **IDT steps ({len(idt_steps)}):**")
            for i, step in enumerate(idt_steps, 1):
                dil = step.get("dilution", "N/A")
                conc = step.get("concentration", "N/A")
                raw = step.get("raw", "")
                lines.append(f"  {i}. Dilution: {dil}, Concentration: {conc}  *(raw: `{raw}`)*")
        else:
            lines.append("- **IDT steps:** *(none found)*")

        lines.append("")

    return "\n".join(lines)


def main():
    print("=" * 70)
    print("Extracting Missing Drug Data from Reference Spreadsheets")
    print("=" * 70)

    # Get existing pages
    existing_pages = get_existing_pages()
    print(f"\nExisting drug pages ({len(existing_pages)}):")
    for p in sorted(existing_pages):
        print(f"  - {p}")

    # Extract all spreadsheet data
    print("\n" + "-" * 70)
    print("Parsing spreadsheets...")
    print("-" * 70)
    all_drugs = extract_all_spreadsheet_data()
    print(f"\nTotal unique drugs found in spreadsheets: {len(all_drugs)}")

    # Find missing drugs
    missing_drugs = {}
    for key, entries in sorted(all_drugs.items()):
        if key not in existing_pages:
            missing_drugs[key] = entries

    print(f"\nDrugs in spreadsheets but MISSING from docs/drugs/: {len(missing_drugs)}")
    for key in sorted(missing_drugs.keys()):
        entries = missing_drugs[key]
        names = [e['original_name'] for e in entries]
        print(f"  - {key} (original names: {', '.join(names)})")

    # Generate output markdown
    output_lines = []
    output_lines.append("# Missing Drugs Data")
    output_lines.append("")
    output_lines.append(f"Generated from reference spreadsheets. These drugs exist in the spreadsheets but do NOT have corresponding pages in `docs/drugs/`.")
    output_lines.append("")
    output_lines.append(f"**Total missing drugs:** {len(missing_drugs)}")
    output_lines.append("")
    output_lines.append("---")
    output_lines.append("")

    for key in sorted(missing_drugs.keys()):
        entries = missing_drugs[key]
        output_lines.append(format_drug_entry(key, entries))
        output_lines.append("---")
        output_lines.append("")

    # Also list all drugs found in spreadsheets for reference
    output_lines.append("## All Drugs Found in Spreadsheets")
    output_lines.append("")
    output_lines.append(f"Total: {len(all_drugs)}")
    output_lines.append("")
    output_lines.append("| # | Normalized Name | Original Name(s) | Has Page |")
    output_lines.append("|---|----------------|-------------------|----------|")
    for i, (key, entries) in enumerate(sorted(all_drugs.items()), 1):
        names = ", ".join(set(e['original_name'] for e in entries))
        has_page = "✅" if key in existing_pages else "❌"
        output_lines.append(f"| {i} | `{key}` | {names} | {has_page} |")

    output_text = "\n".join(output_lines)

    # Save to file
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, 'w') as f:
        f.write(output_text)

    print(f"\n✅ Output saved to: {OUTPUT_FILE}")
    print(f"\n{'=' * 70}")
    print("SUMMARY")
    print(f"{'=' * 70}")
    print(f"Existing drug pages: {len(existing_pages)}")
    print(f"Drugs in spreadsheets: {len(all_drugs)}")
    print(f"Missing drugs (no page): {len(missing_drugs)}")
    print(f"Coverage: {len(all_drugs) - len(missing_drugs)}/{len(all_drugs)} "
          f"({100 * (len(all_drugs) - len(missing_drugs)) / max(len(all_drugs), 1):.1f}%)")


if __name__ == "__main__":
    main()
